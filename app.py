from flask import Flask, render_template, request, redirect, send_file, flash

from config import DevelopmentConfig
from forms import DateForm
import pandas as pd
import numpy as np
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config.from_object(DevelopmentConfig)
app.secret_key = 'your_secret_key'
df = None
input_file_name = None

@app.route('/')
def index():
    form = DateForm()
    return render_template('index.html', form=form)

@app.route('/load', methods=['POST'])
def load_data():
    global df, input_file_name
    if 'file' not in request.files:
        flash('No file part')
        return redirect('/')
    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect('/')
    try:
        df = pd.read_csv(file, encoding='latin1')
        df['DateTime'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
        df.dropna(subset=['DateTime'], inplace=True)
        df['Print Pages'] = pd.to_numeric(df['Print Pages'], errors='coerce').fillna(0).astype(int)
        input_file_name, _ = os.path.splitext(file.filename)
        flash(f'File Name :  {input_file_name}')
        flash('Upload Successful!')

    except Exception as e:
        flash(f'Failed to load data: {e}')
    return redirect('/')

@app.route('/export', methods=['POST'])
def export_data():
    global df, input_file_name
    form = DateForm()
    if form.validate_on_submit():
        start_date = form.start_date.data
        end_date = form.end_date.data

        if start_date > end_date:
            flash('Start date must be before end date')
            return redirect('/')

        if df is None:
            flash('Data not loaded')
            return redirect('/')

        try:
            # Convert start_date and end_date to datetime
            start_date = datetime.combine(start_date, datetime.min.time())
            end_date = datetime.combine(end_date, datetime.min.time())

            # Debugging statements
            print("Start Date:", start_date)
            print("End Date:", end_date)
            print("DataFrame head:", df.head())

            filtered_df = df[(df['DateTime'] >= start_date) & (df['DateTime'] <= end_date)].copy()
            filtered_df.loc[filtered_df['User Name'] == '-', 'User Name'] = 'Photo Copy'
            filtered_df.loc[filtered_df['User Name'] == '?', 'User Name'] = np.nan
            filtered_df.loc[filtered_df['User Name'] == 'support', 'User Name'] = 'pronto'

            total_pages = filtered_df['Print Pages'].sum()

            pages_per_day = filtered_df.groupby(filtered_df['DateTime'].dt.strftime('%d/%m/%Y'))['Print Pages'].sum().reset_index()
            pages_per_day.columns = ['Period', 'Pages']

            pages_per_week = filtered_df.groupby(filtered_df['DateTime'].dt.to_period('W').apply(lambda x: x.start_time))['Print Pages'].sum().reset_index()
            pages_per_week.columns = ['Period', 'Pages']
            pages_per_week['Period'] = pages_per_week['Period'].apply(lambda x: f"{x.strftime('%d/%m/%Y')} - {(x + pd.Timedelta(days=6)).strftime('%d/%m/%Y')}")

            pages_per_month = filtered_df.groupby(filtered_df['DateTime'].dt.to_period('M').apply(lambda x: x.start_time))['Print Pages'].sum().reset_index()
            pages_per_month.columns = ['Period', 'Pages']
            pages_per_month['Period'] = pages_per_month['Period'].apply(lambda x: f"{x.strftime('%d/%m/%Y')} - {(x + pd.offsets.MonthEnd(1)).strftime('%d/%m/%Y')}")

            pages_per_user = filtered_df.groupby('User Name')['Print Pages'].sum().reset_index()
            pages_per_user.columns = ['Username', 'Pages']

            daily_df = pages_per_day
            weekly_df = pages_per_week
            monthly_df = pages_per_month
            user_df = pages_per_user

            combined_df = pd.concat([daily_df, weekly_df, monthly_df, user_df], axis=1)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                combined_df.to_excel(writer, sheet_name='Printer Report', index=False, startrow=2, header=False)
                workbook = writer.book
                sheet = workbook['Printer Report']

                sheet.merge_cells('A1:B1')
                sheet['A1'] = 'Daily'
                sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

                sheet['A2'] = 'Period'
                sheet['B2'] = 'Pages'

                sheet.merge_cells('C1:D1')
                sheet['C1'] = 'Weekly'
                sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

                sheet['C2'] = 'Period'
                sheet['D2'] = 'Pages'

                sheet.merge_cells('E1:F1')
                sheet['E1'] = 'Monthly'
                sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')

                sheet['E2'] = 'Period'
                sheet['F2'] = 'Pages'

                sheet.merge_cells('G1:H1')
                sheet['G1'] = 'Users'
                sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')

                sheet['G2'] = 'Username'
                sheet['H2'] = 'Pages'

                sheet.merge_cells('I1:I2')
                sheet['I1'] = 'Total Pages'
                sheet['I1'].alignment = Alignment(horizontal='center', vertical='center')

                sheet['I3'] = total_pages
                sheet['I3'].alignment = Alignment(horizontal='center', vertical='center')

                for col in range(1, 10):
                    column_letter = get_column_letter(col)
                    if column_letter in ['C', 'E']:
                        sheet.column_dimensions[column_letter].width = 170 / 7.4
                    else:
                        sheet.column_dimensions[column_letter].width = 125 / 7.4

                for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=9):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                header_fill2 = PatternFill(start_color="807D7D", end_color="807D7D", fill_type="solid")
                header_font = Font(bold=True)

                for col in range(1, 10):
                    cell = sheet.cell(row=1, column=col)
                    cell.fill = header_fill2
                    cell.font = header_font
                    cell = sheet.cell(row=2, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

            output.seek(0)
            return send_file(output,
                             download_name=f"{input_file_name}_{start_date.strftime('%d-%m-%Y')}_{end_date.strftime('%d-%m-%Y')}.xlsx",
                             as_attachment=True)


        except Exception as e:
            flash(f'Failed to export data: {e}')
            return redirect('/')

    flash('Invalid form submission')
    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True, port=5001)
